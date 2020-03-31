from collections import defaultdict
from datetime import datetime
import json
from openpyxl import Workbook
import os
import requests
import time
from typing import Any, Callable, Dict, List, Set, Tuple
import pandas as pd
import numpy as np

from const import (
    IU_QUALITY_MAPPING,
    DIFFERENCES_SEARCH_TERMS,
    BIG_TESTING_ROUND_SEARCH_TERMS,
    TOP_100_USER_SEARCHES,
    RESULTS_SHEET,
    MASTER_RESULTS_FILE,
    EXCLUDED_TERMS,
)

from get_profile_url import get_profile_url
from multi_relevance_comparison import process_explain
from stemming_search_term_query import (
    make_search_term_body,
    make_search_term_body_no_low_news,
    make_search_term_with_taxonomy_body,
)
from build_rnmi_dataset import read_master
from competitors_query import make_default_search_term_body, make_competitors_search_term_body_v2


URL = "http://elasticsearch-prd.cbinsights.com:9200/org-read/_analyze"
HEADERS = {"Content-Type": "application/json"}

LIGHT_ENGLISH_STEMMER = "light_english"
MINIMAL_ENGLISH_STEMMER = "minimal_english"
STEM_DIR = "stemming_results"

CUTOFF_RESULTS_FILE = "/Users/emarcey/python/elasticsearch_playground/cutoff_results.xlsx"
OVERRIDE_RESULTS_FILE = "/Users/emarcey/python/elasticsearch_playground/override_results.xlsx"
TAXONOMY_RESULTS_FILE = "/Users/emarcey/python/elasticsearch_playground/taxonomy_results.xlsx"


def make_compound_key(row: Dict[str, Any]) -> str:
    if row["id_org"] == "N/A":
        return f"{row['search_term']}||||N/A"
    if isinstance(row["id_org"], float):
        return f"{row['search_term']}||||{int(row['id_org'])}"
    return f"{row['search_term']}||||{row['id_org']}"


def check_good_bad_results(good_bad_results, row):
    if row["id_org"] == "N/A":
        return {"quality": "MISSING"}
    return good_bad_results[make_compound_key(row)]


def get_good_bad_results() -> Dict[str, str]:
    master_results = read_master(MASTER_RESULTS_FILE, RESULTS_SHEET)

    match_results = defaultdict(dict)
    for _, row in master_results.iterrows():
        match_results[make_compound_key(row)] = {
            "quality": row["quality"],
            "reason_cat": row["reason_cat"],
            "reason": row["reason"],
        }

    return match_results


def make_search_url(index: str) -> str:
    return f"http://elasticsearch-prd.cbinsights.com:9200/{index}/_search"


def compare_result_ranks(
    phrases: List[str], all_hits_1: List[Dict[str, Any]], name_1: str, all_hits_2: List[Dict[str, Any]], name_2: str
) -> Dict[str, List[Dict[str, Any]]]:

    all_results: Dict[str, List[Dict[str, Any]]] = {}
    all_summaries: Dict[str, Dict[str, Any]] = {}
    differences: Set[str] = set()

    for phrase in phrases:

        hits_1 = all_hits_1[phrase]
        hits_2 = all_hits_2[phrase]
        num_hits_1 = len(hits_1)
        num_hits_2 = len(hits_2)

        summary = {
            "num_hits_1": num_hits_1,
            "num_hits_2": num_hits_2,
            "num_same_rank": 0,
            "num_same_orgs": 0,
        }

        hits_1_ranks: Dict[str, int] = {}
        j = 0
        while j < num_hits_1:
            hit_1 = hits_1[j]
            hits_1_ranks[hit_1["_id"]] = j + 1
            j += 1

        hits_2_ranks: Dict[str, int] = {}
        j = 0
        while j < num_hits_2:
            hit_2 = hits_2[j]
            hits_2_ranks[hit_2["_id"]] = j + 1
            j += 1

        min_hits = min(num_hits_1, num_hits_2)
        max_hits = max(num_hits_1, num_hits_2)

        results: List[Dict[str, Any]] = []

        i = 0
        while i < min_hits:
            hit_1 = hits_1[i]
            hit_2 = hits_2[i]

            hits_match = hit_1["_id"] == hit_2["_id"]
            if hits_match:
                summary["num_same_rank"] += 1

            other_rank_1 = hits_2_ranks.get(hit_1["_id"], "N/A")
            if other_rank_1 != "N/A":
                summary["num_same_orgs"] += 1

            other_rank_2 = hits_1_ranks.get(hit_2["_id"], "N/A")
            # if other_rank_2 != "N/A":
            # summary["num_same_orgs"] += 1

            results.append(
                {
                    "rank": i + 1,
                    name_1: hit_1,
                    name_2: hit_2,
                    "hits_match": hits_match,
                    f"{name_2} rank": other_rank_1,
                    f"{name_1} rank": other_rank_2,
                }
            )
            i += 1

        if num_hits_1 > min_hits:
            while i < num_hits_1:
                hit_1 = hits_1[i]
                results.append(
                    {
                        "rank": i + 1,
                        name_1: hit_1,
                        "hits_match": False,
                        f"{name_2} rank": "N/A",
                        f"{name_1} rank": i + 1,
                    }
                )
                i += 1

        if num_hits_2 > min_hits:
            while i < num_hits_2:
                hit_2 = hits_2[i]
                results.append(
                    {
                        "rank": i + 1,
                        name_2: hit_2,
                        "hits_match": False,
                        f"{name_2} rank": i + 1,
                        f"{name_1} rank": "N/A",
                    }
                )
                i += 1

        all_results[phrase] = results

        summary["prop_same_rank"] = "N/A"
        summary["prop_same_orgs"] = "N/A"
        if max_hits > 0:
            summary["prop_same_rank"] = round((summary["num_same_rank"] / max_hits), 2)
            summary["prop_same_orgs"] = round((summary["num_same_orgs"] / max_hits), 2)
            if summary["prop_same_orgs"] != 1:
                differences.add(phrase)
        all_summaries[phrase] = summary

    return all_results, all_summaries, differences


def format_ranked_hit(
    phrase: str, ranked_hit: Dict[str, Any], name_1: str, name_2: str, good_bad_results, line_num
) -> List[str]:
    hit_1 = ranked_hit.get(name_1, {})
    hit_1["search_term"] = phrase
    hit_1["id_org"] = hit_1.get("_id", "N/A")
    hit_2 = ranked_hit.get(name_2, {})
    hit_2["search_term"] = phrase
    hit_2["id_org"] = hit_2.get("_id", "N/A")

    hit_1_source = hit_1.get("_source", {})
    hit_2_source = hit_2.get("_source", {})

    hit_1_good_bad_results = check_good_bad_results(good_bad_results, hit_1)
    hit_2_good_bad_results = check_good_bad_results(good_bad_results, hit_2)

    return [
        phrase,
        ranked_hit["rank"],
        hit_1.get("_id", "N/A"),
        hit_1_source.get("org_name", "N/A"),
        hit_1_good_bad_results.get("quality", ""),
        hit_1_good_bad_results.get("reason_cat", ""),
        hit_1_good_bad_results.get("reason", ""),
        hit_2.get("_id", "N/A"),
        hit_2_source.get("org_name", "N/A"),
        hit_2_good_bad_results.get("quality", ""),
        hit_2_good_bad_results.get("reason_cat", ""),
        hit_2_good_bad_results.get("reason", ""),
        ranked_hit["hits_match"],
        ranked_hit[f"{name_2} rank"],
        ranked_hit[f"{name_1} rank"],
        ranked_hit["rank"] <= 10,
        f'=OR(EQ(E{line_num}, ""), EQ(J{line_num}, ""))',
    ]


def write_result_ranks(
    ws, all_results: Dict[str, List[Dict[str, Any]]], name_1: str, name_2: str, good_bad_results
) -> None:
    ws.append(
        [
            "search term",
            "rank",
            f"{name_1} id_org",
            f"{name_1} org_name",
            f"{name_1} match quality",
            f"{name_1} Reason Cat",
            f"{name_1} Reason",
            f"{name_2} id_org",
            f"{name_2} org_name",
            f"{name_2} match quality",
            f"{name_2} Reason Cat",
            f"{name_2} Reason",
            "hits match",
            f"{name_2} rank",
            f"{name_1} rank",
            "is_top_10",
            "has_empty",
        ]
    )
    line_num = 2
    for phrase, results in all_results.items():
        for result in results:
            ws.append(format_ranked_hit(phrase, result, name_1, name_2, good_bad_results, line_num))
            line_num += 1


def compare_result_ids(
    phrases: List[str], hits_1: List[Any], hits_2: List[Any]
) -> Dict[str, Tuple[Dict[str, Any], Dict[str, Any]]]:
    results: Dict[str, Tuple[Dict[str, Any], Dict[str, Any]]] = {}
    for phrase in phrases:
        hit_1 = hits_1[phrase]
        hit_2 = hits_2[phrase]
        docs_1 = {hit["_id"]: hit for hit in hit_1}
        docs_2 = {hit["_id"]: hit for hit in hit_2}

        missing_1 = {}
        for idDocument, hit in docs_1.items():
            if idDocument not in docs_2:
                missing_1[idDocument] = hit

        missing_2 = {}
        for idDocument, hit in docs_2.items():
            if idDocument not in docs_1:
                missing_2[idDocument] = hit

        results[phrase] = (missing_1, missing_2)
    return results


def get_hits_from_es(response: Dict[str, Any]) -> List[Dict[str, Any]]:
    return response["hits"]["hits"]


def query_es(index: str, phrases: List[str], size: int, query_builder) -> Dict[str, List[Dict[str, Any]]]:
    url = make_search_url(index)
    results: Dict[str, List[Dict[str, Any]]] = {}
    print(f"Querying index: {index}")
    for phrase in phrases:
        print(f"\tQuerying phrase: {phrase}")
        request = query_builder(phrase, size)
        resp = requests.post(url=url, data=json.dumps(request), headers=HEADERS)
        if resp.status_code != 200:
            raise Exception(f"Bad response: {resp.json()} {request}")
        results[phrase] = get_hits_from_es(resp.json())

    return results


def _make_count_if_all(column: str, line_num: int) -> str:
    return f"=COUNTIF(results!{column}:{column}, A{line_num})"


def _make_count_if_diff(column: str, line_num: int, differences: Set[str]) -> str:
    if not differences:
        return 0
    count_if_list = []

    for diff in differences:
        count_if_list.append(f'COUNTIFS(results!{column}:{column}, A{line_num}, results!A:A, "{diff}")')

    return "=" + " + ".join(count_if_list)


def _make_difference_calc(line_num: int) -> str:
    return f"=ROUND(B{line_num}-C{line_num}, 2)"


def _make_pct_calc(col: str, quality_map: Dict[str, int], quality: str, total_line_num: int) -> str:
    quality_line_num = quality_map[quality]
    return f"=ROUND({col}{quality_line_num}/{col}{total_line_num}*100, 2)"


def _make_strict_calc(col: str, quality_map: Dict[str, int], total_line_num: int) -> str:
    r_line_num = quality_map["R"]
    return f"=ROUND({col}{r_line_num}/{col}{total_line_num}*100, 2)"


def _make_loose_calc(col: str, quality_map: Dict[str, int], total_line_num: int) -> str:
    r_line_num = quality_map["R"]
    n_line_num = quality_map["N"]
    return f"=ROUND(({col}{r_line_num}+{col}{n_line_num})/{col}{total_line_num}*100, 2)"


def _make_permissive_calc(col: str, quality_map: Dict[str, int], total_line_num: int) -> str:
    r_line_num = quality_map["R"]
    n_line_num = quality_map["N"]
    m_line_num = quality_map["M"]
    return f"=ROUND(({col}{r_line_num}+{col}{n_line_num}+{col}{m_line_num})/{col}{total_line_num}*100, 2)"


def build_compare_results(ws, differences: Set[str], old_name: str, new_name: str, quality_types: List[str]):

    header = ["", old_name, new_name, "Difference"]
    ws_lines = [header]

    count_if_all_old = "=COUNTIF(results!E:E, $A${0})"
    count_if_all_new = "=COUNTIF(results!H:H, $A${0})"
    difference = "=$B${0}-$C${0}"

    old_col = "E"
    new_col = "J"

    line_num = 2
    all_quality_line_num = {}
    min_all_line_num = line_num
    max_all_line_num = line_num
    for quality_type in quality_types:
        ws_lines.append(
            [
                quality_type,
                _make_count_if_all(old_col, line_num),
                _make_count_if_all(new_col, line_num),
                _make_difference_calc(line_num),
            ]
        )
        all_quality_line_num[quality_type] = line_num
        max_all_line_num = line_num
        line_num += 1

    ws_lines.append(
        [
            "Total",
            f"=SUM(B{min_all_line_num}:B{max_all_line_num})",
            f"=SUM(C{min_all_line_num}:C{max_all_line_num})",
            _make_difference_calc(line_num),
        ]
    )
    all_total_line_num = line_num
    line_num += 1

    ws_lines.append(
        [
            "Strict",
            _make_strict_calc("B", all_quality_line_num, all_total_line_num),
            _make_strict_calc("C", all_quality_line_num, all_total_line_num),
            _make_difference_calc(line_num),
        ]
    )
    line_num += 1
    ws_lines.append(
        [
            "Loose",
            _make_loose_calc("B", all_quality_line_num, all_total_line_num),
            _make_loose_calc("C", all_quality_line_num, all_total_line_num),
            _make_difference_calc(line_num),
        ]
    )
    line_num += 1

    ws_lines.append(
        [
            "Permissive",
            _make_permissive_calc("B", all_quality_line_num, all_total_line_num),
            _make_permissive_calc("C", all_quality_line_num, all_total_line_num),
            _make_difference_calc(line_num),
        ]
    )
    line_num += 1

    ws_lines.append([])
    line_num += 1
    ws_lines.append(["Differences Only"])
    line_num += 1
    ws_lines.append(header)
    line_num += 1

    diff_quality_line_num = {}
    min_all_line_num = line_num
    max_all_line_num = line_num
    for quality_type in quality_types:
        ws_lines.append(
            [
                quality_type,
                _make_count_if_diff(old_col, line_num, differences),
                _make_count_if_diff(new_col, line_num, differences),
                _make_difference_calc(line_num),
            ]
        )
        diff_quality_line_num[quality_type] = line_num
        max_all_line_num = line_num
        line_num += 1

    ws_lines.append(
        [
            "Total",
            f"=SUM(B{min_all_line_num}:B{max_all_line_num})",
            f"=SUM(C{min_all_line_num}:C{max_all_line_num})",
            _make_difference_calc(line_num),
        ]
    )
    diff_total_line_num = line_num
    line_num += 1

    ws_lines.append(
        [
            "Strict",
            _make_strict_calc("B", diff_quality_line_num, diff_total_line_num),
            _make_strict_calc("C", diff_quality_line_num, diff_total_line_num),
            _make_difference_calc(line_num),
        ]
    )
    line_num += 1
    ws_lines.append(
        [
            "Loose",
            _make_loose_calc("B", diff_quality_line_num, diff_total_line_num),
            _make_loose_calc("C", diff_quality_line_num, diff_total_line_num),
            _make_difference_calc(line_num),
        ]
    )
    line_num += 1

    ws_lines.append(
        [
            "Permissive",
            _make_permissive_calc("B", diff_quality_line_num, diff_total_line_num),
            _make_permissive_calc("C", diff_quality_line_num, diff_total_line_num),
            _make_difference_calc(line_num),
        ]
    )
    line_num += 1

    for line in ws_lines:
        ws.append(line)


def _wrap_in_quotes(s: str) -> str:
    return '"' + s + '"'


def _make_rnmi_score(col: str, line_num: int, quality_types: List[str]) -> str:

    rows = []
    i = 2
    for quality_type in quality_types:
        rows.append(
            f"COUNTIFS(results!{col}:{col}, {_wrap_in_quotes(quality_type)}, results!A:A, A{line_num}) * rnmi_penalties!$B${i}"
        )
        i += 1
    return "=" + " + ".join(rows)


def _make_ranked_rnmi_score(col: str, line_num: int, quality_types: List[str]) -> str:
    rows = []
    i = 2
    for quality_type in quality_types:
        rows.append(
            f"COUNTIFS(results!{col}:{col}, {_wrap_in_quotes(quality_type)}, results!A:A, A{line_num}, results!P:P, TRUE) * rnmi_penalties!$B${i}*2"
        )
        rows.append(
            f"COUNTIFS(results!{col}:{col}, {_wrap_in_quotes(quality_type)}, results!A:A, A{line_num}, results!P:P, FALSE) * rnmi_penalties!$B${i}"
        )
        i += 1
    return "=" + " + ".join(rows)


def _make_num_reviewed(col: str, line_num: int, quality_types: List[str]) -> str:
    rows = []
    for quality_type in quality_types:
        rows.append(f"COUNTIFS(results!{col}:{col}, {_wrap_in_quotes(quality_type)}, results!A:A, A{line_num})")
    return "=" + " + ".join(rows)


def write_summary(ws, phrase_summaries, name_1, name_2, quality_types):
    line_num = 1
    ws.append(
        [
            "phrase",
            f"{name_1} Num Results",
            f"{name_2} Num Results",
            "Num same rank",
            "Prop Same Rank",
            "Num same orgs",
            "Prop Same Orgs",
            f"{name_1} Num Reviewed",
            f"{name_2} Num Reviewed",
            f"{name_1} RNMI Score",
            f"{name_2} RNMI Score",
            f"{name_1} Ranked RNMI Score",
            f"{name_2} Ranked RNMI Score",
        ]
    )
    line_num += 1
    # for overall
    line_num += 1

    rows = []
    overall_summary = {"num_hits_1": 0, "num_hits_2": 0, "num_same_rank": 0, "num_same_orgs": 0}
    for phrase, summary in phrase_summaries.items():
        rows.append(
            [
                phrase,
                summary["num_hits_1"],
                summary["num_hits_2"],
                summary["num_same_rank"],
                summary["prop_same_rank"],
                summary["num_same_orgs"],
                summary["prop_same_orgs"],
                _make_num_reviewed("E", line_num, quality_types),
                _make_num_reviewed("J", line_num, quality_types),
                _make_rnmi_score("E", line_num, quality_types),
                _make_rnmi_score("J", line_num, quality_types),
                _make_ranked_rnmi_score("E", line_num, quality_types),
                _make_ranked_rnmi_score("J", line_num, quality_types),
            ]
        )
        overall_summary["num_hits_1"] += summary["num_hits_1"]
        overall_summary["num_hits_2"] += summary["num_hits_2"]
        overall_summary["num_same_rank"] += summary["num_same_rank"]
        overall_summary["num_same_orgs"] += summary["num_same_orgs"]
        line_num += 1

    overall_summary["prop_same_rank"] = round((overall_summary["num_same_rank"] / overall_summary["num_hits_1"]), 2)
    overall_summary["prop_same_orgs"] = round((overall_summary["num_same_orgs"] / overall_summary["num_hits_1"]), 2)

    ws.append(
        [
            "Overall",
            overall_summary["num_hits_1"],
            overall_summary["num_hits_2"],
            overall_summary["num_same_rank"],
            overall_summary["prop_same_rank"],
            overall_summary["num_same_orgs"],
            overall_summary["prop_same_orgs"],
            "=SUM(H3:H1000)",
            "=SUM(I3:I1000)",
            "=SUM(J3:J1000)",
            "=SUM(K3:K1000)",
            "=SUM(L3:L1000)",
            "=SUM(M3:M1000)",
        ]
    )
    for row in rows:
        ws.append(row)


def build_explain(ws, results: Dict[str, List[Dict[str, Any]]]) -> None:
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["C"].width = 20

    ws.append(
        [
            "search term",
            "ranking",
            "Company Name",
            "description",
            "company url",
            "profile url",
            "IU Quality",
            "score",
            "description_score",
            "news_score",
            "cxn_name_score",
            "funding_multiplier",
            "num_cxn_multiplier",
            "name/url score",
            "competitor score",
            "investor score",
            "expert collections",
            "number of expert collections",
            "org_total_equity_funding",
            "tier1_news_score",
            "tier2_news_score",
            "tier3_news_score",
            "tier4_news_score",
            "tier5_news_score",
            "tier6_news_score",
            "tier7_news_score",
            "tier8_news_score",
            "tier9_news_score",
            "tier10_news_score",
            "article_count_multiplier",
            "id_cbi_entity",
            "profile views",
            "mosaic_overall",
            "last funding date",
            "news article count",
            "org_num_fundings",
            "org_num_deals",
            "number of investors",
        ]
    )

    for search_term, result in results.items():
        id_row = 1
        for hit in result:
            explain = hit["_explanation"]
            data = hit["_source"]

            hit_quality = IU_QUALITY_MAPPING.get(search_term, {})

            (
                description_score,
                news_score,
                cxn_name_score,
                funding_multiplier,
                num_cxn_multiplier,
                tier1_news,
                tier2_news,
                tier3_news,
                tier4_news,
                tier5_news,
                tier6_news,
                tier7_news,
                tier8_news,
                tier9_news,
                tier10_news,
                article_count_multiplier,
                name_url_score,
                competitor_score,
                investor_score,
            ) = process_explain(explain, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)

            score = hit["_score"]
            org_name = data.get("org_name", "")
            collections = data.get("expert_collection_names", [])
            profile_url = get_profile_url(data["id_company"])
            quality = hit_quality.get(org_name)

            excel_row = [
                search_term,
                id_row,
                org_name,
                data.get("org_description", ""),
                data.get("org_url", ""),
                profile_url,
                quality,
                score,
                description_score,
                news_score,
                cxn_name_score,
                funding_multiplier,
                num_cxn_multiplier,
                name_url_score,
                competitor_score,
                investor_score,
                str(collections),
                len(collections),
                data.get("org_total_equity_funding", ""),
                tier1_news,
                tier2_news,
                tier3_news,
                tier4_news,
                tier5_news,
                tier6_news,
                tier7_news,
                tier8_news,
                tier9_news,
                tier10_news,
                article_count_multiplier,
                data.get("id_org", ""),
                data.get("org_page_views", ""),
                data.get("org_mosaic_overall", ""),
                data.get("org_last_funding_funding_date", ""),
                data.get("org_news_article_count", ""),
                data.get("org_num_fundings", ""),
                data.get("org_num_deals", ""),
                len(data.get("org_investor", [])),
            ]
            ws.append(excel_row)
            id_row += 1


def make_rnmi_penalties_sheet(ws):
    ws.title = "rnmi_penalties"
    ws.append(["Value", "Penalty"])
    ws.append(["R", 0])
    ws.append(["N", 2])
    ws.append(["M", 4])
    ws.append(["I", 8])


def main() -> None:
    search_terms = TOP_100_USER_SEARCHES
    old_index = "org-read"
    new_index = "org-read"
    size = 25

    old_name = "Current"
    new_name = "New"

    output = []

    quality_types = ["R", "N", "M", "I"]

    wb = Workbook()
    ws = wb.get_active_sheet()

    competitors_score = 3
    investors_score = 2
    name_score = 3

    good_bad_results = read_master(MASTER_RESULTS_FILE, RESULTS_SHEET, EXCLUDED_TERMS)

    old_results = query_es(old_index, search_terms, size, make_default_search_term_body)
    new_results = query_es(
        new_index,
        search_terms,
        size,
        make_competitors_search_term_body_v2(
            name_score=name_score, competitors_score=competitors_score, investors_score=investors_score
        ),
    )

    result_ranks, result_summaries, differences = compare_result_ranks(
        search_terms, old_results, old_name, new_results, new_name
    )
    ws.title = "summaries"
    write_summary(ws, result_summaries, old_name, new_name, quality_types)

    ws = wb.create_sheet()
    ws.title = "compare_results"
    build_compare_results(ws, differences, old_name, new_name, quality_types)

    ws = wb.create_sheet()
    ws.title = RESULTS_SHEET
    write_result_ranks(ws, result_ranks, old_name, new_name, good_bad_results)

    ws = wb.create_sheet()
    make_rnmi_penalties_sheet(ws)

    ws = wb.create_sheet()
    ws.title = f"{old_name} explain"
    build_explain(ws, old_results)

    ws = wb.create_sheet()
    ws.title = f"{new_name} explain"
    build_explain(ws, new_results)

    timestamp = datetime.now().strftime("%Y-%d-%m_%H_%M_%s")
    wb.save(f"{STEM_DIR}/SearchResults_{timestamp}.xlsx")


if __name__ == "__main__":
    main()
